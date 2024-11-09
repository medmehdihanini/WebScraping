from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import pandas as pd
import os
from openpyxl import load_workbook


chrome_service = Service('D:/chromedriver-win64/chromedriver.exe')

chrome_options = Options()
chrome_options.binary_location = "C:/Program Files/Google/Chrome/Application/chrome.exe"

driver = webdriver.Chrome(service=chrome_service, options=chrome_options)

base_url = "{WEB SITE URL }"

all_reviews = []

for page_num in range(40, 49):
    url = base_url + str(page_num)
    driver.get(url)
    time.sleep(3)

    reviews = driver.find_elements(By.CLASS_NAME, "hotel-review-item")
    for review in reviews:
        try:
            author = review.find_element(By.CLASS_NAME, "css-7zzl0z").text
            age_group = review.find_element(By.CLASS_NAME, "css-1ombwl1").text
            review_text = review.find_element(By.CLASS_NAME, "css-1mwjmw9").text

            try:
                review_date = review.find_element(By.CLASS_NAME, "css-1wpd2in").text.split('•')[1].strip()  # Mois et année de l'avis
            except:
                review_date = None

            try:
                traveler_type = review.find_element(By.CLASS_NAME, "css-1wpd2in").text.split('•')[0].strip()  # Type de voyageur
            except:
                traveler_type = None

            try:
                stay_duration = review.find_element(By.CLASS_NAME, "css-1wpd2in").text.split('•')[2].strip()  # Durée du séjour
            except:
                stay_duration = None

            try:
                rating = review.find_element(By.CLASS_NAME, "css-jufmh2").text  # Note de l'avis
            except:
                rating = None

            all_reviews.append({
                "author": author,
                "age_group": age_group,
                "review_date": review_date,
                "traveler_type": traveler_type,
                "stay_duration": stay_duration,
                "rating": rating,
                "review_text": review_text,
            })

        except Exception as e:
            print("Erreur d'extraction : ", e)

    print(f"Page {page_num} extraite.")

driver.quit()

df = pd.DataFrame(all_reviews)

file_path = "yourfilename.xlsx"

if os.path.exists(file_path):
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        book = load_workbook(file_path)
        sheet = book['Sheet1']
        startrow = sheet.max_row
        df.to_excel(writer, index=False, header=False, startrow=startrow)
else:
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)

print("finish extracting data ")